import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import unicodedata
from datetime import datetime

# -------------------------- 配置项 --------------------------
# 你的歌单JSON文件路径（如果和脚本同目录，直接填文件名即可）
JSON_FILE_PATH = "playlist.json"
# 输出的Excel文件路径
OUTPUT_EXCEL_PATH = "网易云歌单导出.xlsx"

# 表格表头（可根据需求调整字段）
HEADERS = [
    "序号", "歌名", "歌手", "专辑", "时长", "发行时间", "歌曲ID"
]

# 样式配置（符合专业视觉规范）
# 表头样式
HEADER_FILL = PatternFill("solid", fgColor="0070C0")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 数据行样式
DATA_FONT = Font(name="微软雅黑", size=10, color="333333")
DATA_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# 交替行背景色（斑马纹）
ZEBRA_FILL_EVEN = PatternFill("solid", fgColor="EBF1F8")
ZEBRA_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")

# 边框样式
THIN_GRAY = Side(style="thin", color="D9D9D9")
BORDER = Border(bottom=THIN_GRAY)

# -------------------------- 工具函数 --------------------------
def display_width(text):
    """计算文本显示宽度，中文/全角字符算2个宽度，英文算1个"""
    return sum(2 if unicodedata.east_asian_width(c) in ('F','W') else 1 for c in str(text or ''))

def auto_fit_columns(ws, min_w=8, max_w=50, padding=3):
    """自动调整列宽，适配内容长度"""
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        w = max((display_width(c.value) for c in col_cells
                 if not isinstance(c, openpyxl.cell.cell.MergedCell) and c.value is not None), default=0)
        ws.column_dimensions[letter].width = max(min_w, min(w * 1.1 + padding, max_w))

def format_duration(ms):
    """将毫秒时长转换为 分:秒 格式"""
    if not ms or not isinstance(ms, (int, float)):
        return ""
    seconds = int(ms / 1000)
    minutes = int(seconds / 60)
    seconds = seconds % 60
    return f"{minutes}:{seconds:02d}"

# -------------------------- 核心逻辑 --------------------------
def parse_playlist_json(json_path):
    """解析网易云歌单JSON，提取歌曲数据（兼容被HTML包裹的JSON）"""
    with open(json_path, "r", encoding="utf-8") as f:
        content = f.read()

    # 去除可能的HTML包裹（如 <html>...<body>{...}</body></html>）
    m = re.search(r"<body>(.*)</body>", content, re.S)
    if m:
        content = m.group(1).strip()

    data = json.loads(content)
    
    # 提取歌单基本信息
    playlist_info = data.get("playlist", {})
    playlist_name = playlist_info.get("name", "未知歌单")
    creator = playlist_info.get("creator", {}).get("nickname", "未知用户")
    song_count = playlist_info.get("trackCount", 0)
    
    # 提取歌曲列表
    tracks = playlist_info.get("tracks", [])
    song_list = []
    
    for idx, track in enumerate(tracks, 1):
        # 提取歌曲核心字段
        song_id = track.get("id", "")
        song_name = track.get("name", "未知歌曲")
        # 处理歌手（多个歌手用/分隔）
        artists = track.get("ar", [])
        artist_names = "/".join([ar.get("name", "") for ar in artists if ar.get("name")])
        # 处理专辑
        album = track.get("al", {})
        album_name = album.get("name", "未知专辑")
        # 处理时长
        duration_ms = track.get("dt", 0)
        duration = format_duration(duration_ms)
        # 处理发行时间
        publish_time = track.get("publishTime", "")
        if publish_time:
            # 转换为YYYY-MM-DD格式
            try:
                publish_time_str = str(publish_time)
                if len(publish_time_str) == 13:
                    # 毫秒级时间戳
                    from datetime import datetime
                    publish_time_dt = datetime.fromtimestamp(int(publish_time_str)/1000)
                    publish_time = publish_time_dt.strftime("%Y-%m-%d")
            except:
                pass
        
        song_list.append([
            idx, song_name, artist_names, album_name, duration, publish_time, song_id
        ])
    
    return {
        "playlist_name": playlist_name,
        "creator": creator,
        "song_count": song_count,
        "song_list": song_list
    }

def create_excel(data, output_path):
    """生成美化后的Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "歌单歌曲明细"
    
    # 写入表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    
    # 写入歌曲数据
    for row_idx, song_data in enumerate(data["song_list"], 2):
        # 交替行背景色
        fill = ZEBRA_FILL_EVEN if (row_idx - 2) % 2 == 0 else ZEBRA_FILL_ODD
        
        for col_idx, value in enumerate(song_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = BORDER
            
            # 设置对齐方式
            if col_idx in [1, 5, 6, 7]:
                # 序号、时长、发行时间、歌曲ID 居中对齐
                cell.alignment = DATA_ALIGN_CENTER
            else:
                # 歌名、歌手、专辑 左对齐
                cell.alignment = DATA_ALIGN_LEFT
    
    # 自动调整列宽
    auto_fit_columns(ws)
    
    # 冻结表头（滚动时表头固定）
    ws.freeze_panes = "A2"
    
    # 新增歌单信息汇总Sheet
    ws_summary = wb.create_sheet("歌单信息汇总")
    summary_data = [
        ["歌单名称", data["playlist_name"]],
        ["创建者", data["creator"]],
        ["歌曲总数", data["song_count"]],
        ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.font = HEADER_FONT if col_idx == 1 else DATA_FONT
            cell.fill = HEADER_FILL if col_idx == 1 else ZEBRA_FILL_EVEN
            cell.alignment = HEADER_ALIGN if col_idx == 1 else DATA_ALIGN_LEFT
            cell.border = BORDER
    
    auto_fit_columns(ws_summary)
    
    # 保存文件
    wb.save(output_path)
    print(f"✅ Excel文件已生成：{output_path}")
    print(f"📊 歌单名称：{data['playlist_name']}")
    print(f"🎵 歌曲总数：{data['song_count']}")

if __name__ == "__main__":
    # 解析JSON数据
    playlist_data = parse_playlist_json(JSON_FILE_PATH)
    # 生成Excel
    create_excel(playlist_data, OUTPUT_EXCEL_PATH)